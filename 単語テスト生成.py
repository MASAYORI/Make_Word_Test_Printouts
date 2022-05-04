import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from glob import glob
import os
import shutil
import string


def input_start_to_end():
    """
    テスト範囲を選択．
    'Start?' => 開始位置．半角数字
    'End?'   => 終了位置．半角数字
    :return: int:start, int:end
    """
    start = int(input('Start?'))
    end = int(input('End?'))
    return start, end


def random_selection(data, start=1, end=100, how=40):
    """
    dataから，howにより指定された数だけランダムにデータを取り出す
    :param data: DataFrame
    :param start: int
    :param end: int
    :param how: int
    :return: DataFrame
    """
    test = data.iloc[start: end].sample(n=how)
    return test


def create_excel(selected_data, start=1, end=100):
    """
    DataFrameをexcel形式に変換
    HACK:tmpdirフォルダが一時的に作られる．formatting関数の終了時に消去される．
    :param selected_data: DataFrame
    :param start: int:ファイル名に使用
    :param end: int:ファイル名に使用
    :return: None
    """
    os.mkdir('tmpdir')
    pd.DataFrame.to_excel(selected_data, 'tmpdir/test_{}-{}.xlsx'.format(start, end), index=None, header=None)


def formatting(database, start, end):
    """
    excelファイルに対しフォーマッティング，解答プリントの解答部分を消去しテストプリントとして保存
    1.各カラムの幅を指定
    2.枠線を引く
    3.テストプリントの作成
    4.tmpdirフォルダの消去
    :param database:
    :param start:
    :param end:
    :return:
    """
    filepaths = glob('tmpdir/test*.xlsx')
    for filepath in filepaths:
        wb = openpyxl.load_workbook(filepath)
        sh = wb.active
        max_col = sh.max_column

        cols = string.ascii_uppercase[:max_col]
        widths = [12]*max_col
        widths[0] = 4.8
        widths[1] = 17
        widths[2] = 65

        sh.move_range('A1:C40', rows=2)
        sh['A1'] = "{}テスト".format(database)
        sh['A1'].font = Font(bold=True)
        sh['C1'] = "{}-{}".format(start, end)
        sh['C1'].font = Font(bold=True)

        width_prefs ={}
        for col, width in zip(cols, widths):
            width_prefs.update({col: width})
        for col_name in width_prefs:
            sh.column_dimensions[col_name].width = width_prefs[col_name]
            for i in range(1, sh.max_row+1):
                sh.row_dimensions[i].height = 18
            wb.save(filepath)

        side = Side(style='thin', color='000000')
        border = Border(left=side, right=side, top=side, bottom=side)
        for row in sh.iter_rows(min_row=3):
            for cell in row:
                cell.border = border
                cell.font = Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(shrinkToFit=True)

        if not os.path.exists('{}_単語テスト'.format(database)):
            os.mkdir(('{}_単語テスト'.format(database)))
        new_filepath = '{}_単語テスト/ans_{}'.format(database, os.path.basename(filepath))
        wb.save(new_filepath)

        for row in sh.iter_rows(min_row=3):
            for cell in row:
                if cell.col_idx == 3:
                    cell.value = None
        wb.save('{}_単語テスト/{}'.format(database, os.path.basename(filepath)))

    shutil.rmtree('tmpdir/')


def main():
    database_dict = {'1': "システム英単語", '2': "ターゲット"}
    database_number = input('数字を入力してください．システム英単語：1 ターゲット：2\n:')
    data = pd.read_csv('{}.csv'.format(database_dict[database_number]))
    print('範囲を選択してください\nどこからどこまでの範囲から選ぶか')
    start, end = input_start_to_end()
    selected_data = random_selection(data=data, start=start, end=end)
    create_excel(selected_data, start=start, end=end)
    formatting(database_dict[database_number], start=start, end=end)


if __name__ == '__main__':
    main()
