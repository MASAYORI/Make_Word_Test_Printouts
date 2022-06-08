import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from glob import glob
import os
import shutil
import string


def random_selection(data, start=1, end=100, how=40):
    """
    dataから，howにより指定された数だけランダムにデータを取り出す
    :param data: DataFrame
    :param start: int
    :param end: int
    :param how: int
    :return: DataFrame
    """

    test = data.iloc[start-1: end].sample(n=how)
    return test


def create_excel(selected_data, start=1, end=100):
    """
    DataFrameをexcel形式に変換
    HACK:tmpdirフォルダが一時的に作られる．formatting関数の終了時に消去される．
    :param selected_data: DataFrame
    :param start: int:ファイル名に使用
    :param end: int:ファイル名に使用
    :return: None
    """
    if os.path.exists('tmpdir'):
        shutil.rmtree('tmpdir/')
    os.mkdir('tmpdir')
    pd.DataFrame.to_excel(selected_data, 'tmpdir/test_{}-{}.xlsx'.format(start, end), index=None, header=None)


def formatting(database, start, end, how):
    """
    excelファイルに対しフォーマッティング，解答プリントの解答部分を消去しテストプリントとして保存
    1.各カラムの幅を指定
    2.枠線を引く
    3.テストプリントの作成
    4.tmpdirフォルダの消去
    """

    filepaths = glob('tmpdir/test*.xlsx')
    for filepath in filepaths:
        wb = openpyxl.load_workbook(filepath)
        sh = wb.active
        max_col = sh.max_column

        cols = string.ascii_uppercase[:max_col]
        widths = [12]*max_col
        widths[0] = 4.8
        widths[1] = 17
        widths[2] = 65

        sh.move_range('A1:C{}'.format(how), rows=2)
        sh['A1'] = "{}テスト".format(database)
        sh['A1'].font = Font(bold=True)
        sh['C1'] = "{}-{}                                            /{}点満点中".format(start, end, how)
        sh['C1'].font = Font(bold=True)

        width_prefs ={}
        for col, width in zip(cols, widths):
            width_prefs.update({col: width})
        for col_name in width_prefs:
            sh.column_dimensions[col_name].width = width_prefs[col_name]
            for i in range(1, sh.max_row+1):
                sh.row_dimensions[i].height = 18
            wb.save(filepath)

        side = Side(style='thin', color='000000')
        border = Border(left=side, right=side, top=side, bottom=side)
        for row in sh.iter_rows(min_row=3):
            for cell in row:
                cell.border = border
                cell.font = Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(shrinkToFit=True)

        if not os.path.exists('{}_単語テスト'.format(database)):
            os.mkdir(('{}_単語テスト'.format(database)))
        new_filepath = '{}_単語テスト/ans_{}'.format(database, os.path.basename(filepath))
        wb.save(new_filepath)

        for row in sh.iter_rows(min_row=3):
            for cell in row:
                if cell.col_idx == 3:
                    cell.value = None
        wb.save('{}_単語テスト/{}'.format(database, os.path.basename(filepath)))

    shutil.rmtree('tmpdir/')


class InputError(Exception):
    pass


def main():
    database_dict = {1: "システム英単語", 2: "ターゲット", 3: "LEAP", 4: "鉄壁", 5: "マドンナ古文", 6: "古文単語315"}
    database_number = int(input('数字を入力してください．システム英単語：1, ターゲット：2, LEAP: 3, \n \
鉄壁: 4, マドンナ古文: 5, 古文単語315: 6\n:'))
    if database_number not in database_dict.keys():
        raise InputError('有効な数字を選択してください．')

    data = pd.read_csv('_word_data/{}.csv'.format(database_dict[database_number]))
    word_count = data.shape[0]

    is_valid = False
    while (is_valid == False):
        print('範囲を選択してください(1-{})\n1)どこから 2)どこまでの範囲で 3)何単語選ぶか'.format(word_count))
        start = int(input('Start?\n:'))
        end = int(input('End?\n:'))

        if start < 1 or start > word_count or end < 1 or end > word_count or start >= end:
            print("範囲選択ミスです．1-{}から選択し直してください．".format(word_count))
            continue
        is_valid = True

    how = int(input('How many?　(Recommendation: 40)\n:'))

    if how > end - start + 1 or how < 1:
        how = end - start + 1
        print('全選択範囲からの出題に設定しました．')
    selected_data = random_selection(data=data, start=start, end=end, how=how)
    create_excel(selected_data, start=start, end=end)
    formatting(database_dict[database_number], start=start, end=end, how=how)


if __name__ == '__main__':
    main()
